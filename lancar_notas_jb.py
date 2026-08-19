#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
LANCAMENTO AUTOMATICO DE NOTAS TOMADAS NO JB CEPIL
Sucesso Contabilidade - Sao Bento do Sul
================================================================================

O QUE ELE FAZ
    1. Le o relatorio PDF (Servico Prestado Analitico) e extrai as notas
    2. Le a planilha de sufixos e acha o codigo/sufixo de cada empresa tomadora
    3. Mostra o plano completo na tela para voce conferir
    4. Depois da sua confirmacao, digita nota por nota dentro do JB Cepil

FILOSOFIA DE SEGURANCA
    O script PARA na hora em qualquer divergencia. Ele nunca "tenta adivinhar".
    Empresa que nao bate exatamente com a planilha nao e lancada - vai pro log.

COMO USAR
    python lancar_notas_jb.py --conferir     (so le e mostra o plano, NAO digita)
    python lancar_notas_jb.py --lancar       (executa de verdade)
    python lancar_notas_jb.py --ancoras      (ajuda a capturar as imagens ancora)

    Retomar um lote interrompido (a NF informada TAMBEM e lancada):
    python lancar_notas_jb.py --lancar --a-partir-de 14928

PARADA DE EMERGENCIA
    Jogue o mouse para o canto superior esquerdo da tela a qualquer momento.
================================================================================
"""

import argparse
import calendar
import csv
import datetime as dt
import os
import re
import sys
import time
import unicodedata

# ==============================================================================
# CONFIGURACAO - revise esta secao antes do primeiro uso
# ==============================================================================

PASTA_BASE = r"\\SUCESSOSERVER\dados\SCAN\AARQUIVOS TRANSITÓRIOS\ALE\Automacoes\Lançamento de nota tomada no jb\Relatório e planilha"

ARQ_PLANILHA = os.path.join(PASTA_BASE, "SUFIXOS NOTAS SUCESSO.xlsx")
ARQ_RELATORIO = os.path.join(PASTA_BASE, "Report.pdf")
PASTA_LOGS = os.path.join(PASTA_BASE, "logs")
PASTA_ANCORAS = os.path.join(PASTA_BASE, "ancoras")

# --- Valores fixos do lancamento (conforme o manual) --------------------------
PACOTE = "3103"

# Empresas com filiais pedem o numero do estabelecimento na tela de pesquisa.
# Digitar o 1 explicitamente (em vez de dar Enter em branco) garante que a
# matriz seja escolhida tanto nas empresas com filial quanto nas sem.
ESTABELECIMENTO = "1"

MODELO = "04"
SERIE = "1"
CNPJ_PRESTADOR = "27969182000160"   # Sucessocont - sempre a prestadora
CFOP = "1933"
ATIVIDADE = "1719"
CODIGO_APOS_ATIVIDADE = "9"

# No inicio do mes o JB costuma nao estar configurado para contabilizar no
# mes corrente, empresa por empresa. Pacote especial que abre o mes: digita
# 3000, 2 enters, a data final do mes contabil, Ctrl+Enter e Esc. Repetir
# isso quando o mes ja esta aberto nao causa problema, entao o script faz
# essa configuracao em toda empresa, assim que entra nela.
PACOTE_CONFIG_MES = "3000"

# --- Ritmo da digitacao -------------------------------------------------------
# Aumente se o JB estiver lento ou a rede pesada. Comece folgado.
PAUSA_TECLA = 0.20       # entre cada tecla/enter
PAUSA_CAMPO = 0.50       # depois de preencher um campo
PAUSA_TELA = 2.50        # esperando uma tela nova abrir

# Depois de fechar uma janela com Esc, o JB leva um tempo extra para
# devolver o foco ao campo de tras. Digitar antes disso faz o texto se
# perder no vazio - foi o que aconteceu com o pacote 3103 sendo digitado
# logo apos o Esc da configuracao de mes, deixando o campo Pacote vazio.
PAUSA_POS_ESC = 1.50
TIMEOUT_TELA = 20.0      # desiste de esperar uma tela depois disso (segundos)

# Telas internas (janelas filhas) so podem ser verificadas por imagem, que e
# mais lenta de avaliar. Damos mais folga, ja que o pacote 3103 costuma
# levar alguns segundos carregando.
TIMEOUT_TELA_INTERNA = 45.0

# Velocidade da digitacao, em segundos por caractere.
# NORMAL  - campos de texto comuns (modelo, serie, numero, valor)
# MASCARA - campos com mascara ou dropdown (data, CNPJ). Esses controles
#           processam caractere por caractere e PERDEM digitos se a
#           digitacao for rapida demais - e quando isso acontece o campo
#           fica invalido e engole o Enter, jogando o proximo dado no
#           lugar errado. Se ainda assim falhar, suba para 0.25.
INTERVALO_NORMAL = 0.06
INTERVALO_MASCARA = 0.15

# Pausa extra depois de campos com mascara, para o controle validar o valor
# antes de receber o Enter.
PAUSA_POS_MASCARA = 0.60

# --- Comportamento ------------------------------------------------------------
# ATENCAO: o manual diz "junto do cfop entra o sufixo". Se na pratica o sufixo
# for um campo separado que exige ENTER antes, mude para True e teste com 1 nota.
CFOP_SUFIXO_COM_ENTER = False

# Formato do valor digitado. "512,00" (virgula) ou "512.00" (ponto).
VALOR_COM_VIRGULA = True

# --- Titulos de janela usados na verificacao ----------------------------------
# A comparacao ignora acentos e maiusculas/minusculas, entao nao precisa
# reproduzir a acentuacao exata da barra de titulo.
TITULO_PESQUISA_EMPRESA = "Pesquisa de Empresas/Estabelecimentos"
TITULO_DIGITACAO = "Digitacao de servicos tomados"
TITULO_DUPLICATAS = "Lancamentos de Duplicatas/Titulos"
TITULO_PRINCIPAL = "de aplicacoes da JB Software"

# Janela de busca que o proprio JB abre sozinho se o campo CFOP receber um
# Enter estando vazio (campo de lookup - tem o mesmo comportamento em
# Codigo Produto, Unidade de Medida etc). Nao e uma tela que o script pede
# para abrir - e um sinal de que a navegacao anterior "sobrou" um Enter.
TITULO_PESQUISA_CFOP = "Pesquisa de CFOP"

# Tela que o JB abre sozinho em algumas empresas (as que tem retencao),
# depois da aba de tributacao e antes das duplicatas. Nao ha nada a
# preencher nela pelo processo atual - basta fechar com Esc.
TITULO_RETENCOES = "Lancamentos de Retencoes"

# Janelas que nao servem como alvo de foco (splash, avisos de carregamento)
TITULOS_IGNORADOS = ("splash",)

# Exigir que as imagens ancora sejam encontradas na tela?
# Deixe False ate confirmar com --testar-ancoras que elas funcionam no seu
# monitor. Com False, ancora nao encontrada vira apenas um aviso no log e a
# verificacao continua sendo feita pelo titulo da janela.
ANCORAS_OBRIGATORIAS = False

# --- Blocos da planilha de sufixos --------------------------------------------
# (aba, rotulo, coluna_codigo, coluna_nome, coluna_sufixo, coluna_observacao)
BLOCOS_PLANILHA = [
    ("Planilha1", "Prosper Exitus", 1, 2, 3, 4),
    ("Planilha2", "Sucesso 2000", 1, 2, 3, 4),
    ("Planilha2", "Sucessocont", 8, 9, 10, 11),
]

# ==============================================================================
# INFRAESTRUTURA - log e erros
# ==============================================================================


def sem_acento(texto):
    """Deixa o texto comparavel: sem acento e em minusculo."""
    t = unicodedata.normalize("NFKD", str(texto))
    return t.encode("ascii", "ignore").decode().lower()


class ParadaSeguranca(Exception):
    """Levantada quando algo saiu do esperado. O script para imediatamente."""


class NotaPulada(Exception):
    """
    Levantada quando UMA nota especifica precisa ser abandonada no meio da
    digitacao (ex: empresa exige campo que o script nao sabe preencher).
    Ao contrario de ParadaSeguranca, nao para o lote inteiro - a nota vira
    pendencia e o script segue para a proxima.
    """


class Log:
    """Grava tudo em arquivo e mostra na tela ao mesmo tempo."""

    def __init__(self, pasta):
        os.makedirs(pasta, exist_ok=True)
        carimbo = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        self.caminho = os.path.join(pasta, f"lancamento_{carimbo}.log")
        self.caminho_pendencias = os.path.join(
            pasta, f"pendencias_{carimbo}.csv")
        self.arquivo = open(self.caminho, "w", encoding="utf-8")
        self.pendencias = []

    def _escrever(self, nivel, msg):
        linha = f"{dt.datetime.now():%H:%M:%S} [{nivel}] {msg}"
        print(linha)
        self.arquivo.write(linha + "\n")
        self.arquivo.flush()

    def info(self, msg):
        self._escrever("INFO", msg)

    def aviso(self, msg):
        self._escrever("AVISO", msg)

    def erro(self, msg):
        self._escrever("ERRO", msg)

    def titulo(self, msg):
        linha = "\n" + "=" * 78 + f"\n{msg}\n" + "=" * 78
        print(linha)
        self.arquivo.write(linha + "\n")
        self.arquivo.flush()

    def pendencia(self, numero, empresa, motivo, detalhe=""):
        """Nota que NAO foi lancada - precisa de tratamento manual."""
        self.pendencias.append({
            "numero_nota": numero,
            "empresa": empresa,
            "motivo": motivo,
            "detalhe": detalhe,
        })
        self._escrever("PENDENCIA", f"NF {numero} - {empresa} - {motivo}")

    def gravar_pendencias(self):
        if not self.pendencias:
            return None
        with open(self.caminho_pendencias, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(
                f, fieldnames=["numero_nota", "empresa", "motivo", "detalhe"],
                delimiter=";")
            w.writeheader()
            w.writerows(self.pendencias)
        return self.caminho_pendencias

    def fechar(self):
        self.arquivo.close()


# ==============================================================================
# LEITURA DO RELATORIO PDF
# ==============================================================================

# Linha de nota: numero, data, nome do tomador, CNPJ ou CPF, valor, situacao
RE_NOTA = re.compile(
    r"^(\d{4,7})\s+"
    r"(\d{2}/\d{2}/\d{4})\s+"
    r"(.+?)\s+"
    r"((?:\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})|(?:\d{3}\.\d{3}\.\d{3}-\d{2}))\s+"
    r"R\$([\d\.,]+)\s+"
    r"(\S+)$"
)

# Nome comprido demais: o PDF joga a 1a parte do nome na linha ANTERIOR e a
# 2a parte na linha SEGUINTE, deixando a linha da nota sem nome nenhum.
RE_NOTA_SEM_NOME = re.compile(
    r"^(\d{4,7})\s+"
    r"(\d{2}/\d{2}/\d{4})\s+"
    r"((?:\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})|(?:\d{3}\.\d{3}\.\d{3}-\d{2}))\s+"
    r"R\$([\d\.,]+)\s+"
    r"(\S+)$"
)

# Linha de detalhe logo abaixo de cada nota - deve ser ignorada
RE_LINHA_DETALHE = re.compile(r"^\d{6}\s+\S+\s+")


class Nota:
    def __init__(self, numero, data_emissao, empresa, documento, valor, situacao):
        self.numero = numero
        self.data_emissao = data_emissao      # dd/mm/aaaa
        self.empresa = empresa                 # nome do tomador
        self.documento = documento             # CNPJ ou CPF do tomador
        self.valor = valor                     # texto, ex "1.621,00"
        self.situacao = situacao               # Emitido / Cancelado

        # preenchidos depois, pela planilha
        self.codigo = None
        self.sufixo = None
        self.bloco = None
        self.observacao = None

    def __repr__(self):
        return f"<NF {self.numero} {self.empresa} R${self.valor}>"


def ler_relatorio(caminho, log):
    """Extrai as notas do PDF. Trata nomes quebrados em duas linhas."""
    try:
        import pdfplumber
    except ImportError:
        raise ParadaSeguranca(
            "Biblioteca pdfplumber nao instalada.\n"
            "Rode no prompt de comando:  pip install pdfplumber")

    if not os.path.exists(caminho):
        raise ParadaSeguranca(f"Relatorio nao encontrado:\n  {caminho}")

    notas = []
    total_esperado = None

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            linhas = [l.strip() for l in texto.split("\n")]

            i = 0
            while i < len(linhas):
                linha = linhas[i]

                m_total = re.search(
                    r"Quantidade de documentos para a Compet[eê]ncia:\s*(\d+)", linha)
                if m_total:
                    total_esperado = int(m_total.group(1))

                m = RE_NOTA.match(linha)
                if m:
                    notas.append(Nota(*m.groups()))
                    i += 1
                    continue

                # nome comprido: parte 1 na linha de cima, parte 2 na de baixo
                m_sem_nome = RE_NOTA_SEM_NOME.match(linha)
                if m_sem_nome:
                    parte1 = linhas[i - 1] if i > 0 else ""
                    parte2 = linhas[i + 1] if i + 1 < len(linhas) else ""

                    if RE_LINHA_DETALHE.match(parte1) or RE_NOTA.match(parte1):
                        parte1 = ""
                    if RE_LINHA_DETALHE.match(parte2) or RE_NOTA.match(parte2):
                        parte2 = ""

                    nome = f"{parte1} {parte2}".strip()
                    if not nome:
                        raise ParadaSeguranca(
                            f"Nao consegui montar o nome da empresa da NF "
                            f"{m_sem_nome.group(1)} no PDF.\n"
                            f"  Linha lida: {linha!r}")

                    notas.append(Nota(
                        m_sem_nome.group(1), m_sem_nome.group(2), nome,
                        m_sem_nome.group(3), m_sem_nome.group(4),
                        m_sem_nome.group(5)))
                    i += 1
                    continue
                i += 1

    log.info(f"Notas extraidas do PDF: {len(notas)}")

    if total_esperado is not None:
        log.info(f"Total informado no rodape do relatorio: {total_esperado}")
        if total_esperado != len(notas):
            raise ParadaSeguranca(
                f"Divergencia na leitura do PDF.\n"
                f"  O relatorio diz que tem {total_esperado} documentos, "
                f"mas eu li {len(notas)}.\n"
                f"  Nao vou lancar nada com essa diferenca. "
                f"Verifique se o PDF esta completo.")

    if not notas:
        raise ParadaSeguranca("Nenhuma nota encontrada no relatorio.")

    # numeros repetidos costumam ser nota duplicada no relatorio
    vistos = {}
    for n in notas:
        vistos.setdefault(n.numero, []).append(n)
    for numero, grupo in vistos.items():
        if len(grupo) > 1:
            log.aviso(f"NF {numero} aparece {len(grupo)}x no relatorio")

    return notas


# ==============================================================================
# LEITURA DA PLANILHA DE SUFIXOS
# ==============================================================================

RE_SO_LETRAS = re.compile(r"[^A-Z0-9 ]")
RE_SUFIXOS_SOCIETARIOS = re.compile(r"\b(LTDA|ME|EPP|SA|EIRELI)\b")
RE_ESPACOS = re.compile(r"\s+")


def normalizar(texto):
    """Deixa o nome comparavel: sem acento, sem pontuacao, sem LTDA/ME/EPP."""
    t = unicodedata.normalize("NFKD", str(texto))
    t = t.encode("ascii", "ignore").decode()
    t = t.upper().replace("&", " E ")
    t = RE_SO_LETRAS.sub(" ", t)
    t = RE_SUFIXOS_SOCIETARIOS.sub(" ", t)
    return RE_ESPACOS.sub(" ", t).strip()


def ler_planilha(caminho, log):
    """
    Monta o indice nome_normalizado -> lista de cadastros.
    Se o mesmo nome aparecer com codigos diferentes, marca como ambiguo.
    """
    try:
        import openpyxl
    except ImportError:
        raise ParadaSeguranca(
            "Biblioteca openpyxl nao instalada.\n"
            "Rode no prompt de comando:  pip install openpyxl")

    if not os.path.exists(caminho):
        raise ParadaSeguranca(f"Planilha nao encontrada:\n  {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    indice = {}
    total = 0

    for aba, rotulo, c_cod, c_nome, c_suf, c_obs in BLOCOS_PLANILHA:
        if aba not in wb.sheetnames:
            log.aviso(f"Aba '{aba}' nao existe na planilha - bloco '{rotulo}' ignorado")
            continue

        ws = wb[aba]
        for linha in ws.iter_rows(min_row=1, values_only=False):
            cod = linha[c_cod - 1].value if len(linha) >= c_cod else None
            nome = linha[c_nome - 1].value if len(linha) >= c_nome else None
            suf = linha[c_suf - 1].value if len(linha) >= c_suf else None
            obs = linha[c_obs - 1].value if len(linha) >= c_obs else None

            if not cod or not nome:
                continue
            # pula as linhas de cabecalho que se repetem no meio da planilha
            if not isinstance(cod, int):
                continue
            if str(nome).strip().upper() in ("EMPRESAS", "NOME", "EMPRESA"):
                continue

            chave = normalizar(nome)
            if not chave:
                continue

            indice.setdefault(chave, []).append({
                "bloco": rotulo,
                "codigo": int(cod),
                "nome": str(nome).strip(),
                "sufixo": suf,
                "observacao": str(obs).strip() if obs else None,
            })
            total += 1

    wb.close()
    log.info(f"Cadastros lidos da planilha: {total} ({len(indice)} nomes distintos)")
    return indice


def casar_empresa(nota, indice, log):
    """
    Encontra o cadastro da empresa. SO aceita casamento exato apos normalizar.
    Retorna (ok, motivo, detalhe).
    """
    chave = normalizar(nota.empresa)
    candidatos = indice.get(chave)

    if not candidatos:
        return False, "empresa nao encontrada na planilha", ""

    codigos = {c["codigo"] for c in candidatos}
    if len(codigos) > 1:
        return (False, "nome ambiguo na planilha",
                "codigos: " + ", ".join(str(c) for c in sorted(codigos)))

    # se qualquer cadastro do mesmo nome tem observacao, o lancamento muda
    com_obs = [c for c in candidatos if c["observacao"]]
    if com_obs:
        return (False, "empresa tem observacao na planilha (lancamento diferente)",
                com_obs[0]["observacao"])

    cadastro = candidatos[0]
    sufixo = cadastro["sufixo"]

    if sufixo is None or str(sufixo).strip() == "":
        return False, "sufixo vazio na planilha", ""

    try:
        sufixo_int = int(sufixo)
    except (TypeError, ValueError):
        return False, "sufixo invalido na planilha", f"valor lido: {sufixo!r}"

    if sufixo_int == 0:
        return False, "sufixo 0 (processo diferente)", ""

    nota.codigo = cadastro["codigo"]
    nota.sufixo = str(sufixo_int)
    nota.bloco = cadastro["bloco"]
    nota.observacao = None
    return True, "", ""


# ==============================================================================
# CAMADA DE VERIFICACAO DE TELA
# ==============================================================================


class Verificador:
    """
    Confere se a tela esta como deveria antes de qualquer digitacao.

    Camada 1 - titulo da janela ativa (sempre ligada)
    Camada 2 - imagens ancora (liga sozinha se a pasta 'ancoras' tiver PNGs)
    """

    def __init__(self, log):
        self.log = log
        try:
            import pyautogui
            import pygetwindow
        except ImportError:
            raise ParadaSeguranca(
                "Bibliotecas de automacao nao instaladas.\n"
                "Rode no prompt de comando:\n"
                "  pip install pyautogui pygetwindow pillow opencv-python")

        self.pyautogui = pyautogui
        self.pygetwindow = pygetwindow
        self.pyautogui.FAILSAFE = True
        self.pyautogui.PAUSE = 0

        self.ancoras = {}
        if os.path.isdir(PASTA_ANCORAS):
            from PIL import Image
            for arq in os.listdir(PASTA_ANCORAS):
                if arq.lower().endswith(".png"):
                    caminho = os.path.join(PASTA_ANCORAS, arq)
                    try:
                        # Le com PIL (aceita caminhos com acento no Windows).
                        # O cv2.imread do pyautogui falha em caminhos assim,
                        # entao guardamos a imagem ja carregada, nunca o caminho.
                        imagem = Image.open(caminho).convert("RGB")
                        imagem.load()
                        self.ancoras[os.path.splitext(arq)[0]] = imagem
                    except Exception as e:
                        log.aviso(f"Nao consegui ler a ancora '{arq}': {e}")
        if self.ancoras:
            log.info(f"Ancoras carregadas: {', '.join(sorted(self.ancoras))}")
        else:
            log.aviso("Nenhuma imagem ancora encontrada - "
                      "verificacao apenas por titulo de janela")

        if "codigo_produto" not in self.ancoras:
            log.aviso(
                "Ancora 'codigo_produto' nao capturada - o script NAO vai "
                "conseguir detectar empresas que exigem esse campo extra, e "
                "vai tentar lancar normalmente. Capture com: "
                "python lancar_notas_jb.py --ancoras")

        if "pesquisa_cfop" not in self.ancoras:
            log.aviso(
                "Ancora 'pesquisa_cfop' nao capturada - se a janela "
                "'Pesquisa de CFOP's' abrir sozinha (empresas que exigem "
                "codigo de produto), o script pode nao detectar e fechar "
                "sozinho. Capture com: python lancar_notas_jb.py --ancoras")

        if "retencoes" not in self.ancoras:
            log.aviso(
                "Ancora 'retencoes' nao capturada - se a tela 'Lancamentos "
                "de Retencoes' abrir sozinha (empresas com retencao), o "
                "script pode nao detectar e fechar sozinho. Capture com: "
                "python lancar_notas_jb.py --ancoras")

    # ------------------------------------------------------------------
    def titulo_ativo(self):
        try:
            j = self.pygetwindow.getActiveWindow()
            return j.title if j else ""
        except Exception:
            return ""

    # Titulos que indicam que o foco caiu de volta no terminal, e nao no JB.
    # O caso mais comum: o usuario clicou na janela do console para ver o
    # log rolando, e isso ativa o modo "Selecionar" do PowerShell/CMD, que
    # rouba o foco do teclado ate alguem apertar Esc.
    PADROES_CONSOLE = ("powershell", "prompt de comando", "selecionar",
                       "select ", "cmd.exe", "windows terminal")

    def titulo_e_console(self, titulo):
        t = titulo.lower()
        return any(p in t for p in self.PADROES_CONSOLE)

    def esperar_tela(self, etapa, fragmento_titulo=None, ancora=None,
                     timeout=None):
        """
        Espera a tela ficar pronta aceitando DUAS formas de confirmacao:

          - titulo da janela  -> funciona para janelas independentes
                                 (ex: Pesquisa de Empresas)
          - imagem ancora     -> unica forma de enxergar janelas FILHAS,
                                 que abrem dentro da janela principal do JB
                                 e nao mudam o titulo do Windows
                                 (ex: Digitacao de servicos, Duplicatas)

        Basta uma das duas confirmar para seguir.
        """
        if not fragmento_titulo and not ancora:
            raise ParadaSeguranca(
                f"Etapa '{etapa}' sem forma de verificacao configurada.")

        if ancora and ancora not in self.ancoras:
            if not fragmento_titulo:
                raise ParadaSeguranca(
                    f"Nao consigo verificar a tela da etapa: {etapa}\n"
                    f"  Essa tela abre DENTRO da janela principal do JB, entao "
                    f"o titulo do Windows nao muda.\n"
                    f"  A unica forma de verificar e pela imagem ancora "
                    f"'{ancora}.png', que nao foi encontrada.\n"
                    f"  Rode:  python lancar_notas_jb.py --ancoras\n"
                    f"  e salve o recorte em:\n    {PASTA_ANCORAS}")
            ancora = None

        limite = time.time() + (timeout or TIMEOUT_TELA)
        alvo = sem_acento(fragmento_titulo) if fragmento_titulo else None
        ultimo = ""
        avisou_console = False

        while time.time() < limite:
            ultimo = self.titulo_ativo()

            if alvo and alvo in sem_acento(ultimo):
                return "titulo"

            if ancora and self.procurar_ancora(ancora) is True:
                return "ancora"

            if self.titulo_e_console(ultimo) and not avisou_console:
                self.log.aviso(
                    f"O foco caiu no console ({ultimo!r}). "
                    f"Clique de volta na janela do JB Cepil, ou aperte Esc "
                    f"no console se ele estiver em modo 'Selecionar'.")
                avisou_console = True

            time.sleep(0.25)

        detalhes = []
        if fragmento_titulo:
            detalhes.append(f"  Titulo esperado: {fragmento_titulo!r}")
            detalhes.append(f"  Titulo ativo:    {ultimo!r}")
        if ancora:
            detalhes.append(f"  Imagem ancora nao encontrada na tela: '{ancora}'")
        if self.titulo_e_console(ultimo):
            detalhes.append(
                "  O foco esta no console/terminal, nao no JB Cepil. "
                "Se o console entrou em modo 'Selecionar', isso trava o "
                "teclado ate apertar Esc.")

        raise ParadaSeguranca(
            f"Tela inesperada na etapa: {etapa}\n" + "\n".join(detalhes))

    def esperar_janela(self, fragmento_titulo, etapa, timeout=None):
        """Espera ate a janela ativa conter o fragmento. Se estourar, para tudo."""
        limite = time.time() + (timeout or TIMEOUT_TELA)
        alvo = sem_acento(fragmento_titulo)
        ultimo = ""
        avisou_console = False
        while time.time() < limite:
            ultimo = self.titulo_ativo()
            if alvo in sem_acento(ultimo):
                return True
            if self.titulo_e_console(ultimo) and not avisou_console:
                # Nao para na hora - da 3s pro usuario reagir e clicar de
                # volta no JB, mas ja avisa o que esta acontecendo.
                self.log.aviso(
                    f"O foco caiu no console ({ultimo!r}). "
                    f"Clique de volta na janela do JB Cepil, ou aperte Esc "
                    f"no console se ele estiver em modo 'Selecionar'.")
                avisou_console = True
            time.sleep(0.25)
        raise ParadaSeguranca(
            f"Tela inesperada na etapa: {etapa}\n"
            f"  Esperava uma janela contendo: {fragmento_titulo!r}\n"
            f"  Janela ativa no momento:      {ultimo!r}"
            + ("\n  O foco esta no console/terminal, nao no JB Cepil. "
               "Se o console entrou em modo 'Selecionar', isso trava o "
               "teclado ate apertar Esc - considere desativar o 'Modo de "
               "Selecao Rapida' nas propriedades do console."
               if self.titulo_e_console(ultimo) else ""))

    def procurar_ancora(self, nome, confianca=0.90):
        """
        Retorna True/False/None (None = nao deu para avaliar).
        Procura em todos os monitores, nao so no principal.
        """
        imagem = self.ancoras.get(nome)
        if imagem is None:
            return None
        try:
            # all_screens=True e essencial em setup de dois monitores:
            # sem ele o print pega so o monitor principal.
            tela = self.pyautogui.screenshot(allScreens=True)
        except TypeError:
            tela = self.pyautogui.screenshot()
        except Exception as e:
            self.log.aviso(f"Nao consegui capturar a tela: {e}")
            return None

        try:
            import pyscreeze
            achou = pyscreeze.locate(imagem, tela, confidence=confianca)
            return achou is not None
        except Exception as e:
            # pyautogui/pyscreeze levantam ImageNotFoundException quando
            # simplesmente nao encontram - isso e resposta, nao erro.
            if type(e).__name__ == "ImageNotFoundException":
                return False
            self.log.aviso(f"Nao consegui avaliar a ancora '{nome}': "
                           f"{type(e).__name__}: {e}")
            return None

    def conferir_ancora(self, nome, etapa, confianca=0.90):
        """Se existir uma ancora com esse nome, confere se esta na tela."""
        resultado = self.procurar_ancora(nome, confianca)

        if resultado is None or resultado is True:
            return

        mensagem = (f"A referencia visual '{nome}' nao foi encontrada na tela.")
        if ANCORAS_OBRIGATORIAS:
            raise ParadaSeguranca(
                f"Tela inesperada na etapa: {etapa}\n  {mensagem}")
        self.log.aviso(f"{mensagem} (etapa: {etapa}) - "
                       f"seguindo pela verificacao de titulo de janela")

    def print_da_tela(self, nome_base):
        """Salva um print para voce ver depois o que estava acontecendo."""
        try:
            os.makedirs(PASTA_LOGS, exist_ok=True)
            carimbo = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            destino = os.path.join(PASTA_LOGS, f"erro_{nome_base}_{carimbo}.png")
            self.pyautogui.screenshot(destino)
            return destino
        except Exception as e:
            self.log.aviso(f"Nao consegui salvar o print: {e}")
            return None


# ==============================================================================
# DIGITACAO
# ==============================================================================


class Digitador:
    """Envolve o pyautogui com as pausas certas."""

    def __init__(self, pyautogui, log=None, passo_a_passo=False, fator=1.0):
        self.pg = pyautogui
        self.log = log
        self.passo_a_passo = passo_a_passo
        self.fator = fator

    def _pausa(self, segundos):
        time.sleep(segundos * self.fator)

    def _confirmar(self, descricao):
        if not self.passo_a_passo:
            return
        print(f"\n  >> proximo: {descricao}")
        resposta = input("     Enter para executar, 'p' para parar: ").strip()
        if resposta.lower() == "p":
            raise ParadaSeguranca("Interrompido por voce no modo passo a passo.")

    def texto(self, valor, campo="", mascara=False):
        """
        Digita um valor. Use mascara=True em campos de data, CNPJ e qualquer
        outro com formatacao automatica - eles perdem caracteres se a
        digitacao for rapida.
        """
        valor = str(valor)
        self._confirmar(f"digitar {valor!r} em {campo or 'campo atual'}")

        intervalo = INTERVALO_MASCARA if mascara else INTERVALO_NORMAL
        for caractere in valor:
            self.pg.write(caractere, interval=0)
            self._pausa(intervalo)

        if mascara:
            self._pausa(PAUSA_POS_MASCARA)
        self._pausa(PAUSA_CAMPO)

        if self.log and campo:
            self.log.info(f"      {campo}: {valor}")

    def enter(self, vezes=1, motivo=""):
        self._confirmar(f"{vezes} enter(s)" + (f" - {motivo}" if motivo else ""))
        for _ in range(vezes):
            self.pg.press("enter")
            self._pausa(PAUSA_TECLA)

    def tecla(self, nome, motivo=""):
        self._confirmar(f"tecla {nome}" + (f" - {motivo}" if motivo else ""))
        self.pg.press(nome)
        self._pausa(PAUSA_TECLA)

    def combinacao(self, *teclas, motivo=""):
        self._confirmar("+".join(teclas) + (f" - {motivo}" if motivo else ""))
        self.pg.hotkey(*teclas)
        self._pausa(PAUSA_TECLA)


def formatar_valor(valor_texto):
    """'1.621,00' -> '1621,00' (tira o separador de milhar)."""
    limpo = valor_texto.replace(".", "")
    if not VALOR_COM_VIRGULA:
        limpo = limpo.replace(",", ".")
    return limpo


def formatar_data(data_texto):
    """'17/07/2026' -> '17072026'."""
    return data_texto.replace("/", "")


def ultimo_dia_do_mes_fmt(data_emissao):
    """
    A partir de uma data 'dd/mm/aaaa' da propria nota, devolve o ultimo dia
    real daquele mes no formato ddmmaaaa (ex: '17/07/2026' -> '31072026').
    """
    dia, mes, ano = (int(p) for p in data_emissao.split("/"))
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    return f"{ultimo_dia:02d}{mes:02d}{ano:04d}"


def configurar_mes_da_empresa(nota, ver, dig, log):
    """
    No inicio do mes, o JB frequentemente nao esta configurado para
    contabilizar no mes corrente para a empresa que acabou de ser
    selecionada. Isso se resolve digitando o pacote especial 3000, 2
    enters, a data final do mes contabil, Ctrl+Enter e Esc - sempre no
    campo Pacote da tela principal, antes do pacote normal da nota.

    Repetir essa configuracao quando o mes ja esta aberto nao causa
    problema, entao o script faz isso em toda empresa, assim que entra
    nela - nao ha como confirmar por tela se ja estava configurado ou nao.
    """
    data_fim_mes = ultimo_dia_do_mes_fmt(nota.data_emissao)

    etapa = f"NF {nota.numero} - configurar mes contabil ({data_fim_mes})"

    dig.texto(PACOTE_CONFIG_MES, "Pacote (config. de mes)")
    dig.enter(2)
    dig.texto(data_fim_mes, "Data final da contabilidade", mascara=True)
    dig.combinacao("ctrl", "enter", motivo="confirmar configuracao de mes")
    dig.tecla("esc", "fechar a janela de configuracao de mes")
    time.sleep(PAUSA_TELA)

    # Confere que voltamos para o campo Pacote normal antes de seguir com
    # o pacote de verdade da nota.
    ver.esperar_tela(etapa, fragmento_titulo=TITULO_PRINCIPAL,
                     ancora="campo_pacote")

    # Esse esperar_tela acima passa na hora pelo TITULO da janela principal,
    # que continua o mesmo o tempo todo - ou seja, ele NAO garante que o
    # foco ja voltou para o campo Pacote. Sem essa folga, o pacote da nota
    # e digitado cedo demais e se perde (campo Pacote fica vazio e a tela
    # de digitacao nunca abre).
    time.sleep(PAUSA_POS_ESC)

    log.info(f"      Mes contabil configurado ate {data_fim_mes}")


def _janela_extra_visivel(ver, ancora, titulo):
    """
    Essas telas costumam ser janelas FILHAS (abrem dentro da principal do
    JB), entao o titulo do Windows pode nao mudar - assim como Digitacao de
    servicos e Duplicatas, a deteccao confiavel e pela ancora. Checa tambem
    por titulo, caso a janela abra independente em algum outro layout.
    """
    if ver.procurar_ancora(ancora) is True:
        return True
    return sem_acento(titulo) in sem_acento(ver.titulo_ativo())


def fechar_janela_extra_se_abriu(ver, dig, log, etapa, rotulo, ancora,
                                 titulo, explicacao=""):
    """
    Algumas telas do JB abrem sozinhas no meio do lancamento, sem o script
    pedir. Em todas basta um Esc, mas se o script seguir digitando por cima
    sem perceber, os campos seguintes caem dentro da janela errada e
    bagunçam a nota inteira.

    Fecha com Esc e confirma que sumiu. Se o Esc nao resolver, para o lote -
    seguir digitando nesse estado e mais arriscado que parar e conferir.

    Devolve True se a janela estava aberta e foi fechada.
    """
    if not _janela_extra_visivel(ver, ancora, titulo):
        return False

    log.aviso(f"A janela '{rotulo}' abriu sozinha"
              + (f" ({explicacao})" if explicacao else "")
              + ". Fechando com Esc...")
    dig.tecla("escape", f"fechar '{rotulo}' aberta sem querer")
    time.sleep(PAUSA_TELA)

    if _janela_extra_visivel(ver, ancora, titulo):
        raise ParadaSeguranca(
            f"Etapa: {etapa}\n"
            f"  A janela '{rotulo}' abriu sem querer e o Esc nao conseguiu "
            f"fecha-la.\n"
            f"  Feche manualmente e confira se os campos ja digitados desta "
            f"nota ficaram corretos antes de rodar de novo.")

    # Folga para o JB devolver o foco ao campo de tras antes da proxima
    # digitacao - sem ela o texto seguinte pode se perder no vazio.
    time.sleep(PAUSA_POS_ESC)

    log.info(f"      '{rotulo}' fechada, seguindo normalmente.")
    return True


def fechar_pesquisa_cfop_se_abriu(ver, dig, log, etapa):
    """
    O campo CFOP e um campo de lookup: se ele receber um Enter estando
    vazio, o proprio JB abre sozinho a "Pesquisa de CFOP's". Isso acontece
    nas empresas que exigem codigo de produto, onde o 5o enter da navegacao
    sobra.
    """
    return fechar_janela_extra_se_abriu(
        ver, dig, log, etapa, "Pesquisa de CFOP's", "pesquisa_cfop",
        TITULO_PESQUISA_CFOP,
        "o campo CFOP recebeu um Enter vazio")


def fechar_retencoes_se_abriu(ver, dig, log, etapa):
    """
    Em algumas empresas o JB abre a tela "Lancamentos de Retencoes" depois
    da aba de tributacao, antes das duplicatas. Nao ha nada a preencher
    nela pelo processo atual - basta fechar com Esc e a tela de duplicatas
    aparece na sequencia.
    """
    return fechar_janela_extra_se_abriu(
        ver, dig, log, etapa, "Lancamentos de Retencoes", "retencoes",
        TITULO_RETENCOES,
        "empresa com retencao")


def lancar_uma_nota(nota, vencimento, ver, dig, log):
    """
    Executa a sequencia completa de uma nota.
    Qualquer divergencia levanta ParadaSeguranca e o lote inteiro para.
    """
    etapa = "inicio"

    # --- Tela 1: pesquisa da empresa -------------------------------------
    # Janela independente: da para verificar pelo titulo do Windows.
    etapa = f"NF {nota.numero} - selecionar empresa {nota.codigo}"
    ver.esperar_tela(etapa, fragmento_titulo=TITULO_PESQUISA_EMPRESA,
                     ancora="pesquisa_empresa")

    dig.texto(nota.codigo, "Codigo da empresa")
    dig.enter(1)
    # Empresas com filial param aqui pedindo o estabelecimento. Antes o
    # script dava Enter em branco, o que so funcionava nas empresas sem
    # filial - agora digita o numero da matriz explicitamente.
    dig.texto(ESTABELECIMENTO, "Estabelecimento")
    dig.enter(1)
    time.sleep(PAUSA_TELA)

    # --- Tela 2: campo pacote --------------------------------------------
    # Volta para a janela principal do JB.
    etapa = f"NF {nota.numero} - abrir pacote {PACOTE}"
    ver.esperar_tela(etapa, fragmento_titulo=TITULO_PRINCIPAL,
                     ancora="campo_pacote")

    # No inicio do mes o JB costuma nao estar configurado para contabilizar
    # no mes corrente para esta empresa - configura de novo a cada empresa,
    # ja que repetir nao causa problema quando o mes ja esta aberto.
    configurar_mes_da_empresa(nota, ver, dig, log)

    dig.texto(PACOTE)
    dig.enter(1)
    time.sleep(PAUSA_TELA)

    # --- Tela 3: digitacao de servicos tomados ---------------------------
    # JANELA FILHA: abre dentro da principal, o titulo do Windows NAO muda.
    # So a ancora consegue confirmar que ela abriu.
    etapa = f"NF {nota.numero} - digitacao dos dados basicos"
    ver.esperar_tela(etapa, ancora="digitacao_servicos",
                     timeout=TIMEOUT_TELA_INTERNA)

    dig.texto(MODELO, "Modelo")
    dig.enter(1)
    dig.texto(SERIE, "Serie")
    dig.enter(2)
    dig.texto(nota.numero, "Numero da nota")
    dig.enter(2)
    # Campo de data: tem mascara e dropdown - digitacao lenta obrigatoria
    dig.texto(formatar_data(nota.data_emissao), "Data de emissao", mascara=True)
    dig.enter(2)
    # Campo de CNPJ: tambem tem mascara (pontos, barra e hifen automaticos)
    dig.texto(CNPJ_PRESTADOR, "Fornecedor (CNPJ Sucessocont)", mascara=True)
    dig.enter(3)
    dig.texto(formatar_valor(nota.valor), "Valor contabil")
    dig.enter(2)

    # --- Tributacao: CFOP + sufixo + atividade ---------------------------
    etapa = f"NF {nota.numero} - CFOP {CFOP} e sufixo {nota.sufixo}"
    # 5 enters e o certo para a maioria das empresas. Empresas que exigem
    # "codigo de produto" tem um campo a menos no caminho, entao esse 5o
    # enter sobra e abre sozinho a "Pesquisa de CFOP's" - fechar_pesquisa_
    # cfop_se_abriu detecta isso, fecha com Esc, e o foco volta certinho
    # para o campo CFOP vazio nos dois casos.
    dig.enter(5, "ir para a aba de tributacao")
    # Da tempo da "Pesquisa de CFOP's" renderizar antes de tirar o print -
    # sem essa pausa o screenshot pode ser tirado cedo demais e a ancora
    # nunca bate, mesmo com a janela realmente aberta.
    time.sleep(PAUSA_TELA)
    fechar_pesquisa_cfop_se_abriu(ver, dig, log, etapa)
    dig.texto(CFOP, "CFOP")

    # Algumas empresas abrem um campo extra "codigo de produto" logo apos o
    # CFOP, que o script nao sabe preencher. Em vez de arriscar digitar
    # errado, pulamos so essa nota.
    if ver.procurar_ancora("codigo_produto") is True:
        raise NotaPulada(
            f"Empresa {nota.empresa!r} exige o campo 'codigo de produto' "
            f"apos o CFOP - este script nao sabe preenche-lo.")

    if CFOP_SUFIXO_COM_ENTER:
        dig.enter(1)
    dig.texto(nota.sufixo, "Sufixo")
    dig.enter(1)
    dig.texto(ATIVIDADE, "Atividade")
    dig.enter(1)
    dig.texto(CODIGO_APOS_ATIVIDADE, "Codigo apos atividade")
    dig.enter(6, "abrir a tela de duplicatas")
    time.sleep(PAUSA_TELA)

    # Em algumas empresas o JB abre "Lancamentos de Retencoes" aqui, antes
    # das duplicatas. Basta o Esc - a tela de duplicatas vem na sequencia.
    etapa = f"NF {nota.numero} - tela de retencoes antes das duplicatas"
    if fechar_retencoes_se_abriu(ver, dig, log, etapa):
        time.sleep(PAUSA_TELA)

    # --- Tela 4: duplicatas ----------------------------------------------
    # Tambem e janela filha - verificacao pela ancora.
    etapa = f"NF {nota.numero} - lancamento da duplicata"
    ver.esperar_tela(etapa, ancora="duplicatas",
                     timeout=TIMEOUT_TELA_INTERNA)

    dig.enter(2)
    dig.texto(nota.numero, "Numero do titulo")
    dig.enter(1)
    dig.texto(vencimento, "Vencimento", mascara=True)
    dig.enter(3)
    time.sleep(PAUSA_TELA)

    # --- Volta para a proxima nota ---------------------------------------
    etapa = f"NF {nota.numero} - retorno para a proxima nota"
    dig.tecla("f10", "voltar para a tela inicial")
    time.sleep(PAUSA_TELA)
    ver.esperar_tela(etapa, fragmento_titulo=TITULO_PESQUISA_EMPRESA,
                     ancora="pesquisa_empresa")

    log.info(f"OK  NF {nota.numero} - {nota.empresa} - "
             f"emp {nota.codigo} suf {nota.sufixo} - R$ {nota.valor}")


def sair_para_pesquisa(ver, dig, log, nota, tentativas=5):
    """
    Aperta F10 repetidas vezes ate confirmar a volta para a tela de Pesquisa
    de Empresas. Usado para abandonar com seguranca um lancamento que ficou
    pela metade (ex: NotaPulada). Se nao conseguir confirmar, para o lote
    inteiro - o estado dentro do JB ficou incerto.
    """
    for _ in range(tentativas):
        dig.tecla("f10", "sair do lancamento pulado e voltar para pesquisa de empresas")
        time.sleep(PAUSA_TELA)
        try:
            ver.esperar_tela(
                f"NF {nota.numero} - saida de seguranca apos pular nota",
                fragmento_titulo=TITULO_PESQUISA_EMPRESA,
                ancora="pesquisa_empresa", timeout=3.0)
            log.info(f"      NF {nota.numero} pulada - retornou para "
                     f"Pesquisa de Empresas")
            return
        except ParadaSeguranca:
            continue

    raise ParadaSeguranca(
        f"NF {nota.numero} foi pulada, mas depois de {tentativas} F10 eu nao "
        f"consegui confirmar a volta para a tela de Pesquisa de Empresas.\n"
        f"  O lancamento pode ter ficado pela metade dentro do JB Cepil. "
        f"Confira manualmente antes de rodar de novo.")


# ==============================================================================
# PREPARACAO DO LOTE
# ==============================================================================


def cortar_a_partir_de(notas, numero_inicial, log):
    """
    Descarta as notas anteriores a 'numero_inicial', mantendo a ordem do
    relatorio. Serve para retomar um lote que ja foi parcialmente lancado:
    a nota informada E incluida (ela ainda precisa ser lancada).

    Se o numero nao existir no relatorio, para tudo - lancar a partir do
    lugar errado e pior que nao lancar.
    """
    alvo = str(numero_inicial).strip().lstrip("0")

    for i, nota in enumerate(notas):
        if nota.numero.lstrip("0") == alvo:
            restantes = notas[i:]
            log.aviso(
                f"RETOMADA: comecando na NF {nota.numero} - "
                f"{i} notas anteriores do relatorio foram ignoradas, "
                f"{len(restantes)} seguem para conferencia.")
            return restantes

    raise ParadaSeguranca(
        f"A NF {numero_inicial} nao existe no relatorio.\n"
        f"  Confira o numero - ele precisa ser igual ao que aparece na "
        f"coluna de numero do PDF.\n"
        f"  O relatorio vai da NF {notas[0].numero} ate a "
        f"NF {notas[-1].numero}.")


def preparar_lote(log, numero_inicial=None):
    """Le PDF + planilha, casa as empresas e separa o que da para lancar."""
    log.titulo("LEITURA DOS ARQUIVOS")
    log.info(f"Relatorio: {ARQ_RELATORIO}")
    log.info(f"Planilha:  {ARQ_PLANILHA}")

    notas = ler_relatorio(ARQ_RELATORIO, log)
    indice = ler_planilha(ARQ_PLANILHA, log)

    # O corte vem ANTES da conferencia, para que as pendencias registradas
    # no log e no CSV falem so das notas que este lote realmente cobre.
    if numero_inicial:
        notas = cortar_a_partir_de(notas, numero_inicial, log)

    log.titulo("CONFERENCIA DAS EMPRESAS")
    prontas = []
    for nota in notas:
        if nota.situacao.strip().lower() != "emitido":
            log.pendencia(nota.numero, nota.empresa,
                          f"situacao '{nota.situacao}' (nao e Emitido)")
            continue

        ok, motivo, detalhe = casar_empresa(nota, indice, log)
        if ok:
            prontas.append(nota)
        else:
            log.pendencia(nota.numero, nota.empresa, motivo, detalhe)

    log.info("")
    log.info(f"Prontas para lancar : {len(prontas)}")
    log.info(f"Pendencias          : {len(log.pendencias)}")
    rotulo_total = ("Total considerado   " if numero_inicial
                    else "Total no relatorio  ")
    log.info(f"{rotulo_total}: {len(notas)}")
    return notas, prontas


def mostrar_plano(prontas, vencimento, log):
    log.titulo("PLANO DE LANCAMENTO")
    print(f"{'NF':>8}  {'EMP':>6} {'SUF':>4}  {'EMISSAO':<10} "
          f"{'VALOR':>12}  EMPRESA")
    print("-" * 78)
    for n in prontas:
        print(f"{n.numero:>8}  {n.codigo:>6} {n.sufixo:>4}  "
              f"{n.data_emissao:<10} {n.valor:>12}  {n.empresa[:32]}")
    print("-" * 78)
    print(f"Total: {len(prontas)} notas   |   Vencimento em todas: {vencimento}")


# ==============================================================================
# VALIDACAO DA DATA DE VENCIMENTO
# ==============================================================================


def pedir_vencimento():
    while True:
        bruto = input("\nData de vencimento das duplicatas (ddmmaaaa): ").strip()
        bruto = bruto.replace("/", "").replace("-", "").replace(".", "")
        if not re.fullmatch(r"\d{8}", bruto):
            print("  Formato invalido. Digite 8 numeros, exemplo: 10092026")
            continue
        try:
            data = dt.datetime.strptime(bruto, "%d%m%Y").date()
        except ValueError:
            print("  Data inexistente. Confira o dia e o mes.")
            continue
        hoje = dt.date.today()
        if data < hoje:
            resp = input(f"  {data:%d/%m/%Y} ja passou. Confirma mesmo assim? (s/n) ")
            if resp.strip().lower() != "s":
                continue
        if (data - hoje).days > 400:
            resp = input(f"  {data:%d/%m/%Y} esta muito no futuro. Confirma? (s/n) ")
            if resp.strip().lower() != "s":
                continue
        return bruto


# ==============================================================================
# MODOS DE EXECUCAO
# ==============================================================================


def modo_conferir(log, numero_inicial=None):
    notas, prontas = preparar_lote(log, numero_inicial=numero_inicial)
    mostrar_plano(prontas, "(nao informado no modo conferencia)", log)

    caminho = log.gravar_pendencias()
    if caminho:
        log.titulo("PENDENCIAS - NAO SERAO LANCADAS")
        for p in log.pendencias:
            print(f"  NF {p['numero_nota']:>8}  {p['empresa'][:40]:<40} "
                  f"{p['motivo']}")
        print(f"\nLista completa salva em:\n  {caminho}")


def focar_janela_jb(ver, log, tempo_maximo=15):
    """
    Procura a janela do JB Cepil e traz para frente, sem depender do
    usuario clicar na hora certa. Ignora acentos no titulo e descarta
    janelas auxiliares como o splash de abertura.
    """
    fragmentos = [sem_acento(t) for t in
                  (TITULO_PESQUISA_EMPRESA, TITULO_PRINCIPAL, TITULO_DIGITACAO)]
    limite = time.time() + tempo_maximo

    while time.time() < limite:
        try:
            todas = ver.pygetwindow.getAllWindows()
        except Exception:
            todas = []

        candidatas = []
        for janela in todas:
            titulo = sem_acento(getattr(janela, "title", "") or "")
            if not titulo:
                continue
            if any(ign in titulo for ign in TITULOS_IGNORADOS):
                continue
            if ver.titulo_e_console(titulo):
                continue
            for ordem, frag in enumerate(fragmentos):
                if frag in titulo:
                    candidatas.append((ordem, janela))
                    break

        candidatas.sort(key=lambda x: x[0])
        for _, alvo in candidatas:
            try:
                if getattr(alvo, "isMinimized", False):
                    alvo.restore()
                alvo.activate()
            except Exception:
                # activate() falha as vezes no Windows; tentamos a proxima
                continue
            time.sleep(0.6)
            ativo = ver.titulo_ativo()
            if not ver.titulo_e_console(ativo) and any(
                    f in sem_acento(ativo) for f in fragmentos):
                log.info(f"Janela do JB em foco: {ativo!r}")
                return
        time.sleep(0.5)

    raise ParadaSeguranca(
        "Nao encontrei a janela do JB Cepil aberta (ou nao consegui traze-la "
        "para frente).\n"
        "  Abra o JB Cepil, deixe na tela de Pesquisa de Empresas, "
        "clique nela, e rode de novo.")


def modo_lancar(log, limite=None, devagar=False, passo_a_passo=False,
                numero_inicial=None):
    notas, prontas = preparar_lote(log, numero_inicial=numero_inicial)

    if not prontas:
        log.erro("Nenhuma nota apta a lancar. Nada a fazer.")
        return

    if limite:
        prontas = prontas[:limite]
        log.aviso(f"MODO TESTE: apenas as {len(prontas)} primeiras notas "
                  f"serao lancadas")

    vencimento = pedir_vencimento()
    mostrar_plano(prontas, vencimento, log)

    caminho = log.gravar_pendencias()
    if caminho:
        print(f"\n{len(log.pendencias)} notas NAO serao lancadas. "
              f"Lista salva em:\n  {caminho}")

    log.titulo("ANTES DE CONFIRMAR")
    print("""
  1. O JB Cepil precisa estar ABERTO na tela de Pesquisa de Empresas
     (a mesma tela que aparece ao apertar F10) - o script traz ele para
     frente sozinho, nao precisa clicar
  2. NAO clique na janela do console/PowerShell enquanto o script roda -
     isso rouba o foco do teclado e trava a digitacao no lugar errado
  3. Nao mexa no mouse nem no teclado enquanto o script roda
  4. Para abortar: jogue o mouse no canto superior esquerdo da tela
  5. Na primeira vez, rode com poucas notas e confira o resultado no JB
""")
    resp = input(f"Digite LANCAR para comecar as {len(prontas)} notas: ").strip()
    if resp != "LANCAR":
        log.info("Cancelado pelo usuario. Nada foi digitado.")
        return

    print("\nProcurando a janela do JB Cepil...")
    ver = Verificador(log)
    fator = 3.0 if devagar else 1.0
    if devagar:
        log.aviso("MODO DEVAGAR: todas as pausas triplicadas")
    if passo_a_passo:
        log.aviso("MODO PASSO A PASSO: vou pedir confirmacao antes de cada acao")
        log.aviso("Cuidado: para confirmar voce precisa clicar no console, "
                  "e depois clicar de volta no JB antes de apertar Enter aqui")
    dig = Digitador(ver.pyautogui, log=log, passo_a_passo=passo_a_passo,
                    fator=fator)
    focar_janela_jb(ver, log)

    log.titulo("EXECUCAO")
    lancadas = 0
    puladas = 0
    try:
        for i, nota in enumerate(prontas, start=1):
            log.info(f"--- {i}/{len(prontas)} ---")
            try:
                lancar_uma_nota(nota, vencimento, ver, dig, log)
                lancadas += 1
            except NotaPulada as e:
                log.pendencia(nota.numero, nota.empresa,
                              "exige codigo de produto (nao suportado)", str(e))
                puladas += 1
                sair_para_pesquisa(ver, dig, log, nota)

    except ParadaSeguranca as e:
        print_erro = ver.print_da_tela(f"nf{nota.numero}")
        log.titulo("PARADA DE SEGURANCA")
        log.erro(str(e))
        log.erro(f"Notas lancadas antes de parar: {lancadas}")
        log.erro(f"Parou na nota: NF {nota.numero} - {nota.empresa}")
        if print_erro:
            log.erro(f"Print da tela salvo em: {print_erro}")
        log.erro("Confira o JB Cepil antes de rodar de novo - "
                 "a ultima nota pode ter ficado pela metade.")
        return

    except KeyboardInterrupt:
        log.titulo("INTERROMPIDO PELO USUARIO")
        log.erro(f"Notas lancadas antes da interrupcao: {lancadas}")
        return

    except Exception as e:
        print_erro = ver.print_da_tela("inesperado")
        log.titulo("ERRO INESPERADO")
        log.erro(f"{type(e).__name__}: {e}")
        log.erro(f"Notas lancadas antes do erro: {lancadas}")
        if print_erro:
            log.erro(f"Print da tela salvo em: {print_erro}")
        return

    log.titulo("CONCLUIDO")
    log.info(f"Notas lancadas com sucesso: {lancadas}")
    log.info(f"Notas puladas (codigo de produto): {puladas}")
    log.info(f"Notas pendentes (manuais):  {len(log.pendencias)}")


def modo_testar_ancoras(log):
    """
    Diagnostico: mostra as janelas abertas e testa cada ancora na tela atual.
    Nao digita nada. Use para confirmar que a verificacao visual funciona
    no seu monitor antes de ligar ANCORAS_OBRIGATORIAS.
    """
    ver = Verificador(log)

    log.titulo("JANELAS ABERTAS AGORA")
    try:
        for janela in ver.pygetwindow.getAllWindows():
            titulo = getattr(janela, "title", "") or ""
            if titulo.strip():
                print(f"  {titulo!r}")
    except Exception as e:
        log.aviso(f"Nao consegui listar as janelas: {e}")

    log.titulo("TITULOS QUE O SCRIPT PROCURA")
    for rotulo, alvo in (
            ("Pesquisa de empresas", TITULO_PESQUISA_EMPRESA),
            ("Tela principal", TITULO_PRINCIPAL),
            ("Digitacao de servicos", TITULO_DIGITACAO),
            ("Duplicatas", TITULO_DUPLICATAS)):
        print(f"  {rotulo:<24} -> contendo {alvo!r} (ignorando acentos)")

    log.titulo("TESTE DAS ANCORAS")
    print("Deixe o JB Cepil visivel na tela que voce quer testar.")
    print("Vou capturar em 6 segundos...\n")
    time.sleep(6)

    if not ver.ancoras:
        log.aviso("Nenhuma ancora carregada - nada a testar.")
        return

    for nome in sorted(ver.ancoras):
        resultado = ver.procurar_ancora(nome)
        if resultado is True:
            print(f"  ENCONTRADA     {nome}")
        elif resultado is False:
            print(f"  nao encontrada {nome}  (normal se a tela atual nao e a dela)")
        else:
            print(f"  ERRO ao testar {nome}")

    print(f"\nJanela ativa no momento da captura: {ver.titulo_ativo()!r}")
    print("\nSe a ancora da tela que estava visivel apareceu como ENCONTRADA,")
    print("voce pode mudar ANCORAS_OBRIGATORIAS para True no topo do script.")


def modo_ancoras(log):
    """Ajuda a capturar as imagens de referencia de cada tela."""
    try:
        import pyautogui
    except ImportError:
        raise ParadaSeguranca("Instale primeiro:  pip install pyautogui pillow")

    os.makedirs(PASTA_ANCORAS, exist_ok=True)
    nomes = [
        ("pesquisa_empresa", "tela de Pesquisa de Empresas/Estabelecimentos"),
        ("campo_pacote", "tela principal com o campo Pacote"),
        ("digitacao_servicos", "tela de Digitacao de servicos tomados"),
        ("duplicatas", "tela de Lancamentos de Duplicatas/Titulos"),
        ("codigo_produto", "campo extra 'codigo de produto' que so aparece "
                           "em algumas empresas, logo apos o CFOP"),
        ("pesquisa_cfop", "janela 'Pesquisa de CFOP's' que abre sozinha "
                          "quando o campo CFOP recebe um Enter vazio "
                          "(acontece em empresas que exigem codigo de "
                          "produto) - recorte o titulo ou algo caracteristico "
                          "dela, como o botao 'Pesquisar' ou 'Sair'"),
        ("retencoes", "tela 'Lancamentos de Retencoes' que abre sozinha em "
                      "algumas empresas, depois da tributacao e antes das "
                      "duplicatas - recorte o titulo dela"),
    ]

    print("""
CAPTURA DE ANCORAS
------------------
Para cada tela, deixe o JB Cepil nela e tire um print da TELA INTEIRA.
Depois recorte no Paint apenas um pedaco pequeno e caracteristico
(um rotulo de campo, o titulo da janela) e salve com o nome indicado.

Pedacos pequenos e de alto contraste funcionam melhor que telas inteiras.
""")
    for nome, descricao in nomes:
        destino = os.path.join(PASTA_ANCORAS, f"{nome}.png")
        existe = "JA EXISTE" if os.path.exists(destino) else "faltando"
        print(f"  [{existe:>9}] {nome}.png  ->  {descricao}")

    print(f"\nSalve os recortes em:\n  {PASTA_ANCORAS}")
    resp = input("\nQuer que eu tire um print da tela inteira agora para "
                 "voce recortar? (s/n) ").strip().lower()
    if resp == "s":
        print("Deixe o JB na tela desejada. Capturando em 6 segundos...")
        time.sleep(6)
        carimbo = dt.datetime.now().strftime("%H%M%S")
        destino = os.path.join(PASTA_ANCORAS, f"tela_cheia_{carimbo}.png")
        pyautogui.screenshot(destino)
        print(f"Print salvo em:\n  {destino}")


# ==============================================================================
# PONTO DE ENTRADA
# ==============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Lancamento automatico de notas tomadas no JB Cepil")
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--conferir", action="store_true",
                       help="so le os arquivos e mostra o plano (nao digita nada)")
    grupo.add_argument("--lancar", action="store_true",
                       help="executa o lancamento no JB Cepil")
    grupo.add_argument("--ancoras", action="store_true",
                       help="ajuda a capturar as imagens de referencia das telas")
    grupo.add_argument("--testar-ancoras", action="store_true",
                       dest="testar_ancoras",
                       help="testa as ancoras e lista as janelas abertas "
                            "(diagnostico, nao digita nada)")
    parser.add_argument("--a-partir-de", metavar="NF", dest="a_partir_de",
                        help="comeca da NF informada, ignorando as anteriores "
                             "do relatorio (para retomar um lote interrompido "
                             "- a NF informada TAMBEM e lancada)")
    parser.add_argument("--limite", type=int, metavar="N",
                        help="lanca apenas as N primeiras notas (use --limite 1 "
                             "no primeiro teste)")
    parser.add_argument("--devagar", action="store_true",
                        help="triplica todas as pausas (para JB lento ou "
                             "para acompanhar o que ele faz)")
    parser.add_argument("--passo-a-passo", action="store_true",
                        dest="passo_a_passo",
                        help="pede confirmacao antes de cada campo digitado "
                             "(para descobrir onde a sequencia se perde)")
    args = parser.parse_args()

    log = Log(PASTA_LOGS)
    try:
        if args.conferir:
            modo_conferir(log, numero_inicial=args.a_partir_de)
        elif args.lancar:
            modo_lancar(log, limite=args.limite, devagar=args.devagar,
                        passo_a_passo=args.passo_a_passo,
                        numero_inicial=args.a_partir_de)
        elif args.ancoras:
            modo_ancoras(log)
        elif args.testar_ancoras:
            modo_testar_ancoras(log)
    except ParadaSeguranca as e:
        log.titulo("PARADA DE SEGURANCA")
        log.erro(str(e))
        sys.exit(1)
    finally:
        print(f"\nLog desta execucao:\n  {log.caminho}")
        log.fechar()


if __name__ == "__main__":
    main()
